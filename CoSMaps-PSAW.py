# -*- coding: utf-8 -*-
"""
Created on Fri Jan  1 16:43:34 2021

@author: StormyWeather
"""

import datetime as dt
from psaw import PushshiftAPI
import logging
from openpyxl import load_workbook

api = PushshiftAPI()

file = 'CoSMapsPSAW.xlsx'
start_epoch=int(dt.datetime(2016, 1, 1).timestamp())
sub = 'curseofstrahd'
count = 0

# handler = logging.StreamHandler()
# handler.setLevel(logging.INFO)

# logger = logging.getLogger('psaw')
# logger.setLevel(logging.INFO)
# logger.addHandler(handler)

# Open the spreadsheet, get the last updated value, set the last updated value to now.
wb = load_workbook(filename = file)
ws = wb['CoSMapsPSAW']
start_epoch = ws['G1'].value
print("last updated: {}".format(start_epoch))
ws['G1'] = str(dt.datetime.now())

# Set r as 'new', but if we ran the script already today set it as 'end' to avoid freezing
# due to empty generator. Ridiculous string typing and parsing because it works
# and I'm lazy.
r = 'new'
time_string = str(start_epoch)
time_check = dt.date(int(time_string[0:4]), int(time_string[5:7]), int(time_string[8:10]))
if time_check == dt.date.today():
    r = 'end'
# Create the generator
gen = api.search_submissions(after=start_epoch,
                             sort='asc',
                             subreddit=sub,
                             filter=['url', 'title', 'subreddit', \
                                     'link_flair_text', 'full_link']
                             )

# Loop over the generator
while r != 'end':
    # The next line will freeze if the start_epoch is too recent -- aka, within the same day.
    r = next(gen,'end')
    if r == 'end':
        break
    info = r.d_
    t = str(info['title'])
    # Previously I did this as a csv and had to replace commas in the titles
    # t = t.replace(',',';')
    # print(info['url'])
    try:
        if 'link_flair_text' in info.keys() or \
            'map' in t.lower():
            # print("has link_flair_text")
            if info['link_flair_text'] == 'MAP' or \
                'map' in t.lower():
                print(r)
                print("")
                posted = str(dt.datetime.fromtimestamp(info['created_utc']))
                data = [t, posted, info['url'], info['full_link']]
                # Append the results to the last row of the spreadsheet and increment the count
                ws.append(data)
                count += 1
        else:
            None
    except KeyError:
        None

# Save the spreadsheet
wb.save(file)
print("Rows added: {}".format(count))
